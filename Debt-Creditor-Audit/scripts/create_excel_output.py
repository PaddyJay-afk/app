#!/usr/bin/env python3
"""Create a formatted Excel workbook from the redacted creditor CSV."""
from __future__ import annotations

import argparse
from pathlib import Path

CURRENCY_COLS = {"Balance", "Minimum Payment"}
DATE_COLS = {"Due Date", "Statement Date", "Source Date"}


def money_to_float(value):
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace("$", "").replace(",", ""))
    except ValueError:
        return value


def main() -> None:
    parser = argparse.ArgumentParser(description="Create formatted Excel workbook for parent-friendly debt review.")
    parser.add_argument("--csv", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.read_csv(args.csv).fillna("")
    for col in CURRENCY_COLS & set(df.columns):
        df[col] = df[col].apply(money_to_float)

    out = Path(args.out); out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Master")
        total_debt = sum(v for v in df.get("Balance", []) if isinstance(v, (int, float)))
        total_minimums = sum(v for v in df.get("Minimum Payment", []) if isinstance(v, (int, float)))
        verified = int((df.get("Verification Status", "") == "Verified from statement").sum()) if "Verification Status" in df else 0
        manual = int(df.get("Verification Status", pd.Series(dtype=str)).astype(str).str.contains("Needs manual", case=False, na=False).sum()) if "Verification Status" in df else 0
        summary = pd.DataFrame([
            ["Total known debt", total_debt],
            ["Total minimum monthly payments", total_minimums],
            ["Number of creditors found", len(df)],
            ["Number verified", verified],
            ["Number needing manual confirmation", manual],
        ], columns=["Metric", "Value"])
        summary.to_excel(writer, index=False, sheet_name="Summary")
        instructions = []
        for _, row in df.iterrows():
            instructions.append({
                "Creditor": row.get("Creditor", ""),
                "Plain-English Instructions": row.get("Exact Parent Instructions", ""),
                "Warnings": row.get("Notes", "") or "Use only official creditor phone numbers. Confirm no fee applies before paying by phone.",
            })
        pd.DataFrame(instructions).to_excel(writer, index=False, sheet_name="Parent Instructions")

        wb = writer.book
        fills = {
            "header": PatternFill("solid", fgColor="1F4E78"),
            "low": PatternFill("solid", fgColor="FFF2CC"),
            "manual": PatternFill("solid", fgColor="FCE4D6"),
        }
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fills["header"]
                cell.alignment = Alignment(wrap_text=True)
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                header = str(ws.cell(row=1, column=col_idx).value or "")
                max_len = min(max(len(str(c.value or "")) for c in column_cells) + 2, 60)
                ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len)
                for cell in column_cells[1:]:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if header in CURRENCY_COLS and isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
                    if header in DATE_COLS:
                        cell.number_format = 'mm/dd/yyyy'
            if ws.title == "Master":
                headers = {cell.value: cell.column for cell in ws[1]}
                for row_idx in range(2, ws.max_row + 1):
                    confidence = str(ws.cell(row_idx, headers.get("Confidence", 1)).value or "")
                    status = str(ws.cell(row_idx, headers.get("Verification Status", 1)).value or "")
                    fill = fills["low"] if confidence.lower() == "low" else fills["manual"] if "manual" in status.lower() else None
                    if fill:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row_idx, col_idx).fill = fill
    print(f"Created formatted redacted Excel workbook at {out}.")

if __name__ == "__main__":
    main()
